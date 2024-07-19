# -*- coding: utf-8 -*-
"""
Created on Thu Oct 12 13:57:45 2023

@author: Sarath Babu
"""

import glob
import pandas as pd
import streamlit as st
# from langchain.llms import CTransformers
# from langchain.callbacks.manager import CallbackManager
# from langchain.callbacks.streaming_stdout import StreamingStdOutCallbackHandler 

import google.generativeai as genai

# callback_manager=CallbackManager([StreamingStdOutCallbackHandler()])

# config_python = {'max_new_tokens': 512, 'repetition_penalty': 1.1,'temperature':0}




# from pysqldf import SQLDF
# sqldf = SQLDF(globals())


st.set_page_config(
    page_title="Insight Generation Python",
    page_icon=":bulb:",
)


st.title(":bulb: Insight Generation Python")


def extract_python_code(result):
    import ast
    result = [i.lstrip() for i in result.split("\n")]
    result = [i for i in result if 'read_csv' not in i]
    result = [i for i in result if 'pd.DataFrame' not in i]
    result = list(filter(None,result))
    
    new_result = []
    for i in result:
        try:
            if ast.parse(i):
                new_result.append(i)
        except:
            pass

    new_result = "\n".join(new_result)
    return new_result
        
def execute_python_code(result):
    import sys
    original_stdout = sys.stdout
    with open('output.txt','wt') as f:
        sys.stdout = f
        try:
            exec(result)
        except:
            pass
        sys.stdout = original_stdout

    result_txt=''
    with open('output.txt','r') as f:
        for i in f.readlines():
            result_txt+=i
    return result_txt

# Initialize chat history
if "messages" not in st.session_state:
    st.session_state.messages = []
    
    
session_files=glob.glob("session/*")



file_name = pd.read_excel('session/current.xlsx')['File_Name'].values[0]

@st.cache_data(show_spinner=f"Fetching data from {file_name}")
def read_files(file_name):
    #st.session_state.messages.append({"role": "file_selected", "content": file_selected})
    if 'xlsx' in file_name:
        df1=pd.read_excel(file_name,nrows=1000)
    elif 'csv' in file_name:
        df1=pd.read_csv(file_name,nrows=1000)  
    return df1


df=read_files(file_name)
columns = df.columns.values
columns = [i.replace(" ","_") for i in columns]
df.columns=columns

st.markdown(f'File Selected: {file_name}')

file_name = file_name.split(".")[0]

file_name = file_name.replace(" ","_")



df_context = pd.read_excel('session/current.xlsx',sheet_name='Context_Provider')

context = df_context['Context'].values[0]
    
st.markdown(context)

context = context.replace(file_name,'df')

df_context = pd.read_excel('session/current.xlsx',sheet_name='Additional_Context')

if len(df_context)>0:
    additional_context = "\n".join(df_context['Additional_Context'].values)
    
    additional_context = additional_context.split('\n')
    
    additional_context = list(dict.fromkeys(additional_context).keys())
    
    additional_context = "\n\n".join(additional_context)
    
    context = context + additional_context
    
    st.markdown(additional_context)

st.markdown('Questions and Insights:')

df_insights = pd.read_excel('session/current.xlsx',sheet_name='Insight_Generation_Python')

df_insights_size = len(df_insights)

if len(df_insights)>0:

    df_insights = "\n".join(df_insights['Insights'].values)
    
    df_insights = df_insights.split(';;;')
    
    df_insights = [i for i in df_insights if i!='']
    
    for index,insight_question_answer in enumerate(df_insights):

        insight_question,insight_answer = insight_question_answer.split('***')
        
        insight_question = insight_question.lstrip()
        st.subheader(f'{index+1}. {insight_question}')
        st.code(insight_answer)
        
        # insight_answer = extract_python_code(insight_answer)
        # st.code(insight_answer)
        # insight_answer = execute_python_code(insight_answer)
        # st.markdown(insight_answer)

df_question = pd.read_excel('session/current.xlsx',sheet_name='Generate_Questions')


    
from openpyxl import load_workbook

def add_to_excel_insights_generation(insights):
    if 'Insight_Generation_Python' in pd.ExcelFile('session/current.xlsx').sheet_names:
        print('present')
    else:
        wb2 = load_workbook('session/current.xlsx')
        wb2.create_sheet('Insight_Generation_Python')
        wb2.save('session/current.xlsx')
        
    wb = load_workbook('session/current.xlsx')
    ws = wb['Insight_Generation_Python']  # or wb.active
    ws['A1'] = 'Insights'
    #ws['A2'] = insights
    
    start = df_insights_size+2
    
    ws[f'A{start}']=insights
    
    wb.save('session/current.xlsx')
    
    


@st.cache_resource(show_spinner="Generating Insights from questions.....")
def generate_python(question):
    genai.configure(api_key='')
    model = genai.GenerativeModel(model_name="gemini-1.5-flash-latest")

    prompt = f"""{context}
    Question: Generate a python code and enclose within ``` for the question '{question}'
    Answer:"""
    
    # llm = CTransformers(model="TheBloke/OpenOrca-Platypus2-13B-GGUF", 
    #                     model_file="openorca-platypus2-13b.q4_K_M.gguf", 
    #                     model_type="llama", 
    #                     config=config_python,callback_manager=callback_manager)
    # llm = CTransformers(model='openorca-platypus2-13b.Q4_K_M.gguf', 
    #             config=config_python,
    #             callback_manager=callback_manager)
    # return llm(prompt)
    
    response = model.generate_content(prompt)

    response = response.candidates[0].content.parts[0].text
    
    response = response.replace('```','').replace('python','')
    
    response = response.strip()
    
    return response



# # Use a pipeline as a high-level helper
# from transformers import pipeline

# @st.cache_resource(show_spinner="Generating Insights from questions.....")
# def generate_results_from_dataframe(df_result):
#     pipe = pipeline("table-question-answering", model="google/tapas-base-finetuned-wtq")
#     pipe()



if len(df_question)>df_insights_size:

    questions = "\n".join(df_question['Questions'].values[df_insights_size:])
    
    questions = questions.split('\n')

    insights=[]
    for index,question in enumerate(questions):
        st.subheader(f'{index+1}. {question}')
        
        python_code = generate_python(question)
        st.code(python_code)
        
        # python_code = extract_python_code(python_code)
        # st.code(python_code)
        
        python_code = python_code +';;;'
        insights.append(f"{question}***{python_code}")
        
        
        # python_code = execute_python_code(python_code)
        # st.markdown(python_code)
        
        
        st.divider()
        
    add_to_excel_insights_generation("\n".join(insights))






# Accept user input
if prompt := st.chat_input("Enter your question here"):
    # Add user message to chat history
    st.session_state.messages.append({"role": "user", "content": prompt})
    # Display user message in chat message container
    with st.chat_message("user"):
        st.markdown(prompt)
        
    # full_response = full_response + prompt

    # Display assistant response in chat message container
    with st.chat_message("assistant"):
        message_placeholder = st.empty()
        assistant_response = "User added question"
        message_placeholder.markdown(assistant_response)
    # Add assistant response to chat history
    st.session_state.messages.append({"role": "assistant", "content": assistant_response})
    #st.rerun()
    
import shutil
print('Insight Generation Python: Copy from Session file to Result file')
shutil.copy2('session/current.xlsx',f'results/results_{file_name}.xlsx')
