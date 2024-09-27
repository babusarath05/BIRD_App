# -*- coding: utf-8 -*-
"""
Created on Thu Oct 12 13:57:45 2023

@author: Sarath Babu
"""

import glob
import pandas as pd
import streamlit as st
import sqlite3
import warnings
warnings.filterwarnings('ignore')
# from langchain.llms import CTransformers
# from langchain.callbacks.manager import CallbackManager
# from langchain.callbacks.streaming_stdout import StreamingStdOutCallbackHandler 

import google.generativeai as genai


# callback_manager=CallbackManager([StreamingStdOutCallbackHandler()])

# config_sql = {'max_new_tokens': 512, 'repetition_penalty': 1.1,'temperature':0,'stop':[';']}


# from pysqldf import SQLDF
# sqldf = SQLDF(globals())


def clean_sql_query(query_result):
    if query_result.find('SELECT')!=-1:
        query_result = query_result[query_result.find('SELECT'):]
    elif query_result.find('Select')!=-1:
        query_result = query_result[query_result.find('Select'):]
    else:
        query_result = 'No Select Statement Found'
        
    return query_result

st.set_page_config(
    page_title="Insight Generation SQL",
    page_icon=":bulb:",
)


st.title(":bulb: Insight Generation SQL")

# Initialize chat history
if "messages" not in st.session_state:
    st.session_state.messages = []
    
    
session_files=glob.glob("session/*")



file_name = pd.read_excel('session/current.xlsx')['File_Name'].values[0]

@st.cache_data(show_spinner=f"Fetching data from {file_name}")
def read_files(file_name):
    #st.session_state.messages.append({"role": "file_selected", "content": file_selected})
    if 'xlsx' in file_name:
        df1=pd.read_excel(file_name,nrows=10000)
    elif 'csv' in file_name:
        df1=pd.read_csv(file_name,nrows=10000)  
    return df1


df=read_files(file_name)
columns = df.columns.values
columns = [i.replace(" ","_") for i in columns]
df.columns=columns

# =============================================================================
file = "BIRD.db"
try: 
    conn = sqlite3.connect(file) 
    print(f"Database {file} formed.") 
except: 
    print(f"Database {file} not formed.")

with sqlite3.connect(file) as conn:
    try:
        df.to_sql('df',conn,if_exists='replace',index=False)
        print('df table created')
    except:
        print('df table not created')
        
con = sqlite3.connect("BIRD.db")
cursor = con.cursor()
cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
print(cursor.fetchall())
# =============================================================================


st.markdown(f'File Selected: {file_name}')

file_name = file_name.split(".")[0]

file_name = file_name.replace(" ","_")



df_context = pd.read_excel('session/current.xlsx',sheet_name='Context_Provider')

context = df_context['Context'].values[0]
    
st.markdown(context)

context = context.replace(file_name,'df')


df_context = pd.read_excel('session/current.xlsx',sheet_name='Additional_Context')

if len(df_context)>0:
    additional_context = "\n".join(df_context['Additional_Context'].values)
    
    additional_context = additional_context.split('\n')
    
    additional_context = list(dict.fromkeys(additional_context).keys())
    
    additional_context = "\n\n".join(additional_context)
    
    context = context + additional_context
    
    st.markdown(additional_context)

st.markdown('Questions and Insights:')

df_insights = pd.read_excel('session/current.xlsx',sheet_name='Insight_Generation_SQL')

df_insights_size = len(df_insights)

if len(df_insights)>0:

    df_insights = "\n".join(df_insights['Insights'].values)
    
    df_insights = df_insights.split(';')
    
    df_insights = [i for i in df_insights if i!='\n']
    
    df_insights = [i for i in df_insights if i!='']
    
    
    for index,insight_question_answer in enumerate(df_insights):

        insight_question,insight_answer = insight_question_answer.split('***')
        
        insight_question = insight_question.lstrip()
        st.subheader(f'{index+1}. {insight_question}')
        st.code(insight_answer)
        
        try:
            # df_result = sqldf.execute(insight_answer)
            with sqlite3.connect(file) as conn:
                df_result = pd.read_sql_query(insight_answer, conn)
            st.dataframe(df_result)  
        except Exception as e:
            print(e)
            st.markdown('Sql query incorrect')
    

df_question = pd.read_excel('session/current.xlsx',sheet_name='Generate_Questions')


    
from openpyxl import load_workbook

def add_to_excel_insights_generation(insights):
    if 'Insight_Generation_SQL' in pd.ExcelFile('session/current.xlsx').sheet_names:
        print('present')
    else:
        wb2 = load_workbook('session/current.xlsx')
        wb2.create_sheet('Insight_Generation_SQL')
        wb2.save('session/current.xlsx')
        
    wb = load_workbook('session/current.xlsx')
    ws = wb['Insight_Generation_SQL']  # or wb.active
    ws['A1'] = 'Insights'
    #ws['A2'] = insights
    
    start = df_insights_size+2
    
    ws[f'A{start}']=insights
    
    wb.save('session/current.xlsx')
    
    


@st.cache_resource(show_spinner="Generating Insights from questions.....")
def generate_sql_query(question):
    genai.configure(api_key='')
    model = genai.GenerativeModel(model_name="gemini-1.5-flash-latest")

    prompt = f"""{context}
    Question: Generate a sql query and enclose within ``` for the question '{question}'
    Answer:"""

    # llm = CTransformers(model="TheBloke/OpenOrca-Platypus2-13B-GGUF", 
    #                     model_file="openorca-platypus2-13b.q4_K_M.gguf", 
    #                     model_type="llama", 
    #                     config=config_sql,callback_manager=callback_manager)
    
    # # llm = CTransformers(model='openorca-platypus2-13b.Q4_K_M.gguf', 
    # #             config=config_sql,
    # #             callback_manager=callback_manager)
    # return llm(prompt)

    response = model.generate_content(prompt)

    response = response.candidates[0].content.parts[0].text
    
    response = response.replace('```','').replace('sql','')
    
    response = response.strip()
    
    return response

@st.cache_resource(show_spinner="Generating Insights from questions.....")
def describe_answer(question,query_result,df_result):
    prompt=f"""
    {context}
    This is the Question {question},
    Please refer to Output of SQL Query {df_result} and 
    SQL Query {query_result}
    before answering the question.
    
    Answer to the Question:
    """
    llm = CTransformers(model='openorca-platypus2-13b.Q4_K_M.gguf', 
                        config=config_sql,callback_manager=callback_manager)
    return llm(prompt)

# # Use a pipeline as a high-level helper
# from transformers import pipeline

# @st.cache_resource(show_spinner="Generating Insights from questions.....")
# def generate_results_from_dataframe(df_result):
#     pipe = pipeline("table-question-answering", model="google/tapas-base-finetuned-wtq")
#     pipe()



if len(df_question)>df_insights_size:

    questions = "\n".join(df_question['Questions'].values[df_insights_size:])
    
    questions = questions.split('\n')

    insights=[]
    for index,question in enumerate(questions):
        st.subheader(f'{index+1}. {question}')
        print(f'\n{index+1}. {question}')
        query_result = generate_sql_query(question)
        # query_result = clean_sql_query(query_result)+';'
        st.code(query_result)
        
        try:
            # df_result = sqldf.execute(query_result)
            with sqlite3.connect(file) as conn:
                df_result = pd.read_sql_query(query_result, conn)
            st.dataframe(df_result)  
        except Exception as e:
            print(e)
            st.markdown('Sql query incorrect')
        
        # python_code = generate_python(question)
        # st.code(python_code)
        #st.markdown(describe_answer(question,query_result,df_result))
        
        insights.append(f"{question}***{query_result}")
        st.divider()
        
    add_to_excel_insights_generation("\n".join(insights))


  
import shutil
print('Insight Generation Sql: Copy from Session file to Result file')
shutil.copy2('session/current.xlsx',f'results/results_{file_name}.xlsx')
