# -*- coding: utf-8 -*-
"""
Created on Sat Oct  7 11:24:54 2023

@author: Sarath Babu
"""

import streamlit as st
import time
import glob
import pandas as pd
# from langchain.llms import CTransformers
# from langchain.callbacks.manager import CallbackManager
# from langchain.callbacks.streaming_stdout import StreamingStdOutCallbackHandler 

import google.generativeai as genai

# config = {'max_new_tokens': 512, 'repetition_penalty': 1.1,'temperature':1,'stop':['11']}
# callback_manager=CallbackManager([StreamingStdOutCallbackHandler()])

st.set_page_config(
    page_title="Recommendation Generation",
    page_icon=":scroll:",
)


st.title(":scroll: Recommendation Generation")

# Initialize chat history
if "messages" not in st.session_state:
    st.session_state.messages = []


session_files=glob.glob("session/*")

file_name = pd.read_excel('session/current.xlsx')['File_Name'].values[0]

st.markdown(f'File Selected: {file_name}')

file_name = file_name.split(".")[0]

file_name = file_name.replace(" ","_")

df_context = pd.read_excel('session/current.xlsx',sheet_name='Context_Provider')

context = df_context['Context'].values[0]

st.markdown(context)


df_context = pd.read_excel('session/current.xlsx',sheet_name='Additional_Context')

if len(df_context)>0:
    additional_context = "\n".join(df_context['Additional_Context'].values)
    
    additional_context = additional_context.split('\n')
    
    additional_context = list(dict.fromkeys(additional_context).keys())
    
    additional_context = "\n\n".join(additional_context)
    
    context = context + additional_context
    
    st.markdown(additional_context)

# st.text_area(label="Enter Insights here:")

df_recommendation = pd.read_excel('session/current.xlsx',sheet_name='Recommendation_Generation')

if len(df_recommendation)>0:
    recommendations = "\n".join(df_recommendation['Recommendations'].values)
    
    recommendations = recommendations.split('\n')
    
    recommendations = list(dict.fromkeys(recommendations).keys())
    
    recommendations = "\n\n".join([f'{i+1}. {j}' for i,j in enumerate(recommendations)])
    
    
    st.chat_message('assistant').markdown(recommendations)


from openpyxl import load_workbook

def add_to_excel_generate_recommendations(recommendations):
    if 'Recommendation_Generation' in pd.ExcelFile('session/current.xlsx').sheet_names:
        print('present')
    else:
        wb2 = load_workbook('session/current.xlsx')
        wb2.create_sheet('Recommendation_Generation')
        wb2.save('session/current.xlsx')
        
    wb = load_workbook('session/current.xlsx')
    ws = wb['Recommendation_Generation']  # or wb.active
    ws['A1'] = 'Recommendations'
    
    start = len(df_recommendation)+2
    
    ws[f'A{start}']=recommendations
    # ws['A2'] = questions
    wb.save('session/current.xlsx')


@st.cache_resource(show_spinner="Generating Business Recommendations.....")
def generate_recommendations():
    genai.configure(api_key='')
    model = genai.GenerativeModel(model_name="gemini-1.5-flash-latest")

    prompt=f"""
    {context}
    Question: Generate 10 business recommendations from the context and enclose within ```
    Answer:
    """
    # llm = CTransformers(model="TheBloke/OpenOrca-Platypus2-13B-GGUF", 
    #                     model_file="openorca-platypus2-13b.q4_K_M.gguf", 
    #                     model_type="llama", 
    #                     config=config,callback_manager=callback_manager)
    
    # # llm = CTransformers(model='openorca-platypus2-13b.Q4_K_M.gguf', 
    # #             config=config,
    # #             callback_manager=callback_manager)
    # return llm(prompt)

    response = model.generate_content(prompt)
    
    response = response.candidates[0].content.parts[0].text
    
    response = response.replace("```","")
    
    response = response.replace("**","")
    
    response = response.replace("#","")
    
    return response



# # Display chat messages from history on app rerun
# for message in st.session_state.messages:
#     if (message['role']=='user')|(message['role']=='assistant'):
#         with st.chat_message(message["role"]):
#             st.markdown(message["content"])


if st.button("Generate Business Recommendations"):
    with st.chat_message("user"):
        st.markdown("Generate Business Recommendations")
    # Display assistant response in chat message container
    with st.chat_message("assistant"):
        message_placeholder = st.empty()
        full_response = ""
        #with st.spinner('Generating Responses'):   
        assistant_response = generate_recommendations()
        
        # Simulate stream of response with milliseconds delay
        for chunk in assistant_response.split():
            if chunk in [f"{i}." for i in range(1,11)]:
                full_response+=chunk+" "
            else:
                full_response += chunk + "\n"
            time.sleep(0.05)
            # Add a blinking cursor to simulate typing
            message_placeholder.markdown(full_response + "▌")
        #message_placeholder.markdown(full_response)
        
    #st.session_state.messages.append({"role": "assistant", "content": full_response})
    
    result_business=assistant_response
    result_business=[i.lstrip() for i in result_business.split("\n")]
    result_business=list(filter(None,[(lambda x:x[x.find('.')+2:] if (x.find(".")>=1)&(x.find(".")<=3) else None)(x) for x in result_business]))
    result_business=list(dict.fromkeys(result_business).keys())
    
    
    add_to_excel_generate_recommendations("\n".join(result_business))
    
    st.rerun()

# Accept user input
if prompt := st.chat_input("You can add your business recommendations here"):
    # Add user message to chat history
    st.session_state.messages.append({"role": "user", "content": prompt})
    # Display user message in chat message container
    with st.chat_message("user"):
        st.markdown(prompt)
    
    full_response = ""
    full_response = full_response + prompt

    # Display assistant response in chat message container
    with st.chat_message("assistant"):
        message_placeholder = st.empty()
        assistant_response = "User added question"
        message_placeholder.markdown(assistant_response)
    # Add assistant response to chat history
    st.session_state.messages.append({"role": "assistant", "content": assistant_response})
    #st.rerun()
    
    
import shutil
print('Recommendation Generation: Copy from Session file to Result file')
shutil.copy2('session/current.xlsx',f'results/results_{file_name}.xlsx')
