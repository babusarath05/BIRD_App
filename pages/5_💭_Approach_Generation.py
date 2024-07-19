# -*- coding: utf-8 -*-
"""
Created on Thu Oct 12 13:57:45 2023

@author: Sarath Babu
"""
import glob
import pandas as pd
import streamlit as st
# from langchain.llms import CTransformers
# from langchain.callbacks.manager import CallbackManager
# from langchain.callbacks.streaming_stdout import StreamingStdOutCallbackHandler 

import google.generativeai as genai

# callback_manager=CallbackManager([StreamingStdOutCallbackHandler()])

# config_python = {'max_new_tokens': 512, 'repetition_penalty': 1.1,'temperature':1}

st.set_page_config(
    page_title="Approach Generation",
    page_icon=":thought_balloon:",
)


st.title(":thought_balloon: Approach Generation")

# Initialize chat history
if "messages" not in st.session_state:
    st.session_state.messages = []
    
    
session_files=glob.glob("session/*")


file_name = pd.read_excel('session/current.xlsx')['File_Name'].values[0]

st.markdown(f'File Selected: {file_name}')

file_name = file_name.split(".")[0]

file_name = file_name.replace(" ","_")

df_context = pd.read_excel('session/current.xlsx',sheet_name='Context_Provider')

context = df_context['Context'].values[0]
    
st.markdown(context)


df_context = pd.read_excel('session/current.xlsx',sheet_name='Additional_Context')

if len(df_context)>0:
    additional_context = "\n".join(df_context['Additional_Context'].values)
    
    additional_context = additional_context.split('\n')
    
    additional_context = list(dict.fromkeys(additional_context).keys())
    
    additional_context = "\n\n".join(additional_context)
    
    context = context + additional_context
    
    st.markdown(additional_context)

st.markdown('Questions and Approaches:')


df_question = pd.read_excel('session/current.xlsx',sheet_name='Generate_Questions')

df_approaches = pd.read_excel('session/current.xlsx',sheet_name='Approach_Generation')

df_approaches_size = len(df_approaches)

if len(df_approaches)>0:
    print('Data Present')

    df_approaches = "\n".join(df_approaches['Approaches'].values)
    
    df_approaches = df_approaches.split(';;;')
    
    df_approaches = [i for i in df_approaches if i!='']
    
    for index,insight_question_approach in enumerate(df_approaches):

        insight_question,insight_approach = insight_question_approach.split('***')
        
        insight_question = insight_question.lstrip()
        insight_approach = "\n\n".join([i.lstrip() for i in insight_approach.split('\n')])
        st.subheader(f'{index+1}. {insight_question}')
        st.markdown(insight_approach)

 
from openpyxl import load_workbook

def add_to_excel_approach_generation(approaches):
    if 'Approach_Generation' in pd.ExcelFile('session/current.xlsx').sheet_names:
        print('present')
    else:
        wb2 = load_workbook('session/current.xlsx')
        wb2.create_sheet('Approach_Generation')
        wb2.save('session/current.xlsx')
        
    wb = load_workbook('session/current.xlsx')
    ws = wb['Approach_Generation']  # or wb.active
    ws['A1'] = 'Approaches'
    #ws['A2'] = insights
    
    start = df_approaches_size+2
    
    ws[f'A{start}']=approaches
    
    wb.save('session/current.xlsx')

#Describe only the approaches to solve the question '{question}'
@st.cache_resource(show_spinner="Generating Approaches from questions.....")
def generate_approaches(question):
    genai.configure(api_key='')
    model = genai.GenerativeModel(model_name="gemini-1.5-flash-latest")
    
    prompt=f"""
    {context}
    Question: Describe only the approaches to solve the question '{question}'
    Answer:
    """
    # llm = CTransformers(model="TheBloke/OpenOrca-Platypus2-13B-GGUF", 
    #                     model_file="openorca-platypus2-13b.q4_K_M.gguf", 
    #                     model_type="llama", 
    #                     config=config_python,callback_manager=callback_manager)
    # llm = CTransformers(model='openorca-platypus2-13b.Q4_K_M.gguf', 
    #             config=config_python,
    #             callback_manager=callback_manager)
    # return llm(prompt)
    
    response = model.generate_content(prompt)

    response = response.candidates[0].content.parts[0].text
    
    response = response.strip()
    
    return response



if len(df_question)>df_approaches_size:

    questions = "\n".join(df_question['Questions'].values[df_approaches_size:])
    
    questions = questions.split('\n')

    approaches=[]
    for index,question in enumerate(questions):
        st.subheader(f'{index+1}. {question}')
        
        approach = generate_approaches(question)
        st.markdown(approach)
        approach = approach + ';;;'
        #st.markdown(describe_answer(question,query_result,df_result))
        
        approaches.append(f"{question}***{approach}")
        st.divider()
        
    add_to_excel_approach_generation("\n".join(approaches))


    
import shutil
print('Approach Generation: Copy from Session file to Result file')
shutil.copy2('session/current.xlsx',f'results/results_{file_name}.xlsx')
