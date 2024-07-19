# -*- coding: utf-8 -*-
"""
Created on Tue Oct 10 17:39:43 2023

@author: v-skarunanit
"""

import streamlit as st
import pandas as pd
import glob

st.set_page_config(page_title="Context", page_icon=":clipboard:")

st.title(":clipboard: Context Provider")


# Initialize chat history
if "messages" not in st.session_state:
    st.session_state.messages = []
    
    
session_files=glob.glob("session/*")

df_context = pd.read_excel('session/current.xlsx',engine='openpyxl')

file_selected = df_context['File_Name'].values[0]

file_name = file_selected

@st.cache_data(show_spinner=f"Fetching data from {file_name}")
def read_files(file_name):
    #st.session_state.messages.append({"role": "file_selected", "content": file_selected})
    if 'xlsx' in file_name:
        df=pd.read_excel(file_name,nrows=1,engine='openpyxl')
    elif 'csv' in file_name:
        df=pd.read_csv(file_name,nrows=1)  
    return df


df_context=read_files(file_name)

columns = df_context.columns.values

columns = [i.replace(" ","_") for i in columns]

df_context.columns=columns

file_name = file_name.split(".")[0]

file_name = file_name.replace(" ","_")


df_context = pd.read_excel('session/current.xlsx',sheet_name='Context_Provider')
prompt = df_context['Context'].values[0]

st.markdown(prompt)
#add_to_excel_context(prompt)


from openpyxl import load_workbook

df_additional_context = pd.read_excel('session/current.xlsx',sheet_name='Additional_Context' ,engine='openpyxl')

if len(df_additional_context)>0:
    
    additional_context = "\n".join(df_additional_context['Additional_Context'].values)
    
    additional_context =  additional_context.split("\n")
    
    additional_context = list(dict.fromkeys(additional_context).keys())
    
    additional_context = "\n\n".join(additional_context)
    
    st.markdown(additional_context)


def add_to_excel_additional_context(prompt,file_selected):
    if 'Additional_Context' in pd.ExcelFile('session/current.xlsx').sheet_names:
        print('present')
    else:
        wb2 = load_workbook('session/current.xlsx')
        wb2.create_sheet('Additional_Context')
        wb2.save('session/current.xlsx')
        
    wb = load_workbook('session/current.xlsx')
    ws = wb['Additional_Context']  # or wb.active
    
    start = len(df_additional_context)+2
    
    ws['A1'] = 'Additional_Context'
    ws[f'A{start}'] = prompt
    ws[f'B{start}'] = file_selected
    wb.save('session/current.xlsx')



st.markdown("<u>Ctrl + Enter to confirm</u>",unsafe_allow_html=True)
if additional_context := st.text_area(label="Enter additional context here:"):
    st.markdown(additional_context)
    st.markdown('Now press on Add Prompt')
    
    #st.session_state.messages.append({'role':'prompt','content':f'{additional_context}'})
if st.button('Add Prompt'):
    if additional_context!='':
        add_to_excel_additional_context(additional_context,file_selected)
    st.rerun()
    
import shutil
print('Context Provider: Copy from Session file to Result file')
shutil.copy2('session/current.xlsx',f'results/results_{file_name}.xlsx')
