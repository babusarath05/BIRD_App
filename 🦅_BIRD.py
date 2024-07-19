# -*- coding: utf-8 -*-
"""
Created on Sat Oct  7 11:24:54 2023

@author: Sarath Babu
"""

import streamlit as st
import glob
import pandas as pd
import os
import shutil
import warnings
warnings.filterwarnings('ignore')

st.set_page_config(page_title="BIRD",page_icon=":eagle:")

files=glob.glob("*.csv")+glob.glob("*.xlsx")


st.title(":eagle:BIRD: Buisness Insight & Recommendation Developer")


session_files=glob.glob("session/*")


# if len(session_files)>0:
#     file_selected_first = pd.read_excel('session/current.xlsx')['File_Name'].values[0]
#     files.remove(file_selected_first)
#     files = [file_selected_first]+files
   
    
from openpyxl import load_workbook

def add_to_excel_file(file_selected):
    if 'BIRD' in pd.ExcelFile('session/current.xlsx').sheet_names:
        print('BIRD Sheet present')
    else:
        wb2 = load_workbook('session/current.xlsx')
        wb2.create_sheet('BIRD')
        wb2.save('session/current.xlsx')
        
    wb = load_workbook('session/current.xlsx')
    ws = wb['BIRD']  # or wb.active
    ws['A1'] = 'File_Name'
    ws['A2'] = file_selected

    wb.save('session/current.xlsx')
    
    
def check_file_change(file_selected):
    wb = load_workbook('session/current.xlsx')
    ws = wb['BIRD']  # or wb.active
    
    file_name = file_selected
    file_name = file_name.split(".")[0]
    file_name = file_name.replace(" ","_")
    
    if ws['A2'].value != file_selected:
        print('File Mismatch')
        if f'results_{file_name}.xlsx' in os.listdir('results/'):
            print('Result File Present')
            print('Copy from Result file to Session file')
            shutil.copy2(f'results/results_{file_name}.xlsx','session/current.xlsx')
        else:
            print('Emptying Session file')
            new_sheets = ['Additional_Context',
                          'Generate_Questions',
                          'Insight_Generation_SQL',
                          'Insight_Generation_Python',
                          'Approach_Generation',
                          'Recommendation_Generation']
            for sheet in new_sheets:
                wb = load_workbook('session/current.xlsx')
                if sheet in wb.sheetnames:
                    wb.remove(wb[sheet])
                    print(f'{sheet} removed',end=',')
                wb.save('session/current.xlsx')

                    
def add_to_excel_context(prompt):
    if 'Context_Provider' in pd.ExcelFile('session/current.xlsx').sheet_names:
        print('Context_Provider sheet present')
    else:
        wb2 = load_workbook('session/current.xlsx')
        wb2.create_sheet('Context_Provider')
        wb2.save('session/current.xlsx')
        
    wb = load_workbook('session/current.xlsx')
    ws = wb['Context_Provider']  # or wb.active
    
    ws['A1'] = 'Context'
    ws['A2'] = prompt
    wb.save('session/current.xlsx')

def add_to_excel_new_sheets(sheet,column):
    if sheet in pd.ExcelFile('session/current.xlsx').sheet_names:
        print(f'{sheet} sheet present')
    else:
        wb2 = load_workbook('session/current.xlsx')
        wb2.create_sheet(sheet)
        wb2.save('session/current.xlsx')
        
    wb = load_workbook('session/current.xlsx')
    ws = wb[sheet]  # or wb.active
    
    ws['A1'] = column
    
    wb.save('session/current.xlsx')


file_selected = st.selectbox(
    "Choose your file",
    ['-']+files)

print('\n\nFile_selected:',file_selected)



st.markdown("<u>Please press reset before loading a file</u>",unsafe_allow_html=True)
if st.button('Reset'):
    st.rerun()
    
if len(session_files)==0:
    pd.DataFrame().to_excel('session/current.xlsx',sheet_name='BIRD')
 


@st.cache_data(show_spinner=f"Fetching data from {file_selected}")
def read_files(file_selected):
    #st.session_state.messages.append({"role": "file_selected", "content": file_selected})
    if 'xlsx' in file_selected:
        df=pd.read_excel(file_selected,nrows=100)
    elif 'csv' in file_selected:
        df=pd.read_csv(file_selected,nrows=100)  
    return df

if file_selected != '-':
    
    check_file_change(file_selected)
    
    df=read_files(file_selected)
    columns = df.columns.values
    columns = [i.replace(" ","_") for i in columns]
    df.columns=columns
    print(df.shape)
    
    st.write(df)
    st.write('Displaying only 100 rows')
    
    
    file_name = file_selected#.split(".")[0]
    add_to_excel_file(file_selected)
    
    
    file_name = file_name.split(".")[0]
    
    file_name = file_name.replace(" ","_")
    
    prompt=f"""
    Context:
        
    {file_name} is a dataframe which contains columns such as {",".join(columns)}.
    """
    add_to_excel_context(prompt)
    
    new_sheets = [('Additional_Context','Additional_Context'),
                  ('Generate_Questions','Questions'),
                  ('Insight_Generation_SQL','Insights'),
                  ('Insight_Generation_Python','Insights'),
                  ('Approach_Generation','Approaches'),
                  ('Recommendation_Generation','Recommendations')]
    
    for sheet,column in new_sheets:
        add_to_excel_new_sheets(sheet, column)

    # import shutil
    print('BIRD: Copy from Session file to Result file')
    shutil.copy2('session/current.xlsx',f'results/results_{file_name}.xlsx')
